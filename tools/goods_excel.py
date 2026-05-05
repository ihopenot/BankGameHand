"""goods.yaml + folk.yaml + game.yaml + economy.yaml <-> Excel 双向转换工具。

用法:
    # YAML → Excel（生成 goods.xlsx，含公式自动计算成本/利润/ROI/需求/现金分析）
    python tools/goods_excel.py export

    # Excel → YAML（读取 goods.xlsx 写回 config/goods.yaml 和 config/folk.yaml）
    python tools/goods_excel.py import
"""

import argparse
from pathlib import Path

import yaml
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parent.parent
GOODS_YAML = PROJECT_ROOT / "config" / "goods.yaml"
FOLK_YAML = PROJECT_ROOT / "config" / "folk.yaml"
GAME_YAML = PROJECT_ROOT / "config" / "game.yaml"
ECONOMY_YAML = PROJECT_ROOT / "config" / "economy.yaml"
EXCEL_PATH = PROJECT_ROOT / "goods.xlsx"

# ────────────────────────── helpers ──────────────────────────


def _load_yaml(path):
    with open(path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def _save_yaml(data, path):
    with open(path, "w", encoding="utf-8") as f:
        yaml.dump(data, f, allow_unicode=True, default_flow_style=False, sort_keys=False)


# ────────────────── 样式 ──────────────────

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
SECTION_FONT = Font(bold=True, size=12)
SECTION_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
FORMULA_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
PARAM_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
WARN_FONT = Font(bold=True, color="FF0000")
CHAIN_FILLS = {
    "电子": PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"),
    "纺织": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    "食品": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
}
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

FOLK_NAMES = ["低收入群体", "中等收入群体", "高收入群体"]
# 终端消费品 → 其所在产业链 3 层工厂 (原料, 中间, 终端)
CONSUMER_GOODS = ["手机", "服装", "食品"]

CHAIN_MAP = {
    "硅": "电子", "芯片": "电子", "手机": "电子",
    "硅矿场": "电子", "芯片工厂": "电子", "手机工厂": "电子",
    "硅矿开采": "电子", "芯片制造": "电子", "手机组装": "电子",
    "棉花": "纺织", "布料": "纺织", "服装": "纺织",
    "棉花农场": "纺织", "纺织厂": "纺织", "服装厂": "纺织",
    "棉花种植": "纺织", "布料加工": "纺织", "服装生产": "纺织",
    "小麦": "食品", "面粉": "食品", "食品": "食品",
    "麦田": "食品", "面粉厂": "食品", "食品厂": "食品",
    "小麦种植": "食品", "面粉加工": "食品", "食品生产": "食品",
}


def _style_header(ws, row, col_count):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _style_section(ws, row, col_count, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_count)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = THIN_BORDER


def _style_data(cell, is_formula=False):
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if is_formula:
        cell.fill = FORMULA_FILL


def _style_param(cell):
    cell.fill = PARAM_FILL
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _auto_width(ws):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            val = str(cell.value) if cell.value is not None else ""
            max_len = max(max_len, len(val.encode("gbk", errors="replace")))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 10)


# ────────────────── EXPORT ──────────────────


def export_to_excel():
    goods_data = _load_yaml(GOODS_YAML)
    folk_data = _load_yaml(FOLK_YAML)
    game_data = _load_yaml(GAME_YAML)
    economy_data = _load_yaml(ECONOMY_YAML)

    recipe_map = {r["name"]: r for r in goods_data["recipes"]}
    goods_price_row = {g["name"]: 2 + i for i, g in enumerate(goods_data["goods_types"])}

    wb = Workbook()

    # ═══════════ Sheet 1: 商品类型 ═══════════
    ws1 = wb.active
    ws1.title = "商品类型"
    headers1 = ["名称", "基础价格"]
    ws1.append(headers1)
    _style_header(ws1, 1, len(headers1))
    for g in goods_data["goods_types"]:
        r = ws1.max_row + 1
        ws1.cell(row=r, column=1, value=g["name"])
        ws1.cell(row=r, column=2, value=g["base_price"])
        chain = CHAIN_MAP.get(g["name"])
        for c in range(1, 3):
            cell = ws1.cell(row=r, column=c)
            _style_data(cell)
            if chain in CHAIN_FILLS:
                cell.fill = CHAIN_FILLS[chain]
    _auto_width(ws1)

    # ═══════════ Sheet 2: 配方 ═══════════
    ws2 = wb.create_sheet("配方")
    headers2 = ["名称", "输入商品", "输入数量", "输出商品", "输出数量", "科技品质权重"]
    ws2.append(headers2)
    _style_header(ws2, 1, len(headers2))
    for rec in goods_data["recipes"]:
        r = ws2.max_row + 1
        ws2.cell(row=r, column=1, value=rec["name"])
        ws2.cell(row=r, column=2, value=rec.get("input") or "(无)")
        ws2.cell(row=r, column=3, value=rec["input_quantity"])
        ws2.cell(row=r, column=4, value=rec["output"])
        ws2.cell(row=r, column=5, value=rec["output_quantity"])
        ws2.cell(row=r, column=6, value=rec["tech_quality_weight"])
        chain = CHAIN_MAP.get(rec["name"])
        for c in range(1, 7):
            cell = ws2.cell(row=r, column=c)
            _style_data(cell)
            if chain in CHAIN_FILLS:
                cell.fill = CHAIN_FILLS[chain]
    _auto_width(ws2)

    # ═══════════ Sheet 3: 工厂类型 ═══════════
    ws3 = wb.create_sheet("工厂类型")
    headers3 = ["名称", "配方", "劳动力需求", "建造成本", "维护费", "建造时间"]
    ws3.append(headers3)
    _style_header(ws3, 1, len(headers3))
    for ft in goods_data["factory_types"]:
        r = ws3.max_row + 1
        ws3.cell(row=r, column=1, value=ft["name"])
        ws3.cell(row=r, column=2, value=ft["recipe"])
        ws3.cell(row=r, column=3, value=ft["labor_demand"])
        ws3.cell(row=r, column=4, value=ft["build_cost"])
        ws3.cell(row=r, column=5, value=ft["maintenance_cost"])
        ws3.cell(row=r, column=6, value=ft["build_time"])
        chain = CHAIN_MAP.get(ft["name"])
        for c in range(1, 7):
            cell = ws3.cell(row=r, column=c)
            _style_data(cell)
            if chain in CHAIN_FILLS:
                cell.fill = CHAIN_FILLS[chain]
    _auto_width(ws3)

    # ═══════════ Sheet 4: 居民配置 ═══════════
    ws_folk = wb.create_sheet("居民配置")
    #  A          B        C        D       E          F              G
    # 群体名称  人口  品质权重  品牌权重  价格权重  劳动参与率  人均劳动力
    #  H             I            J            K             L            M
    # 食品人均需求  食品敏感度  服装人均需求  服装敏感度  手机人均需求  手机敏感度
    folk_headers = [
        "群体名称", "人口", "品质权重", "品牌权重", "价格权重",
        "劳动参与率", "人均劳动力",
        "食品-人均需求", "食品-敏感度",
        "服装-人均需求", "服装-敏感度",
        "手机-人均需求", "手机-敏感度",
    ]
    ws_folk.append(folk_headers)
    _style_header(ws_folk, 1, len(folk_headers))

    # 居民商品顺序: 食品、服装、手机（与 folk.yaml base_demands 中的 key 对应）
    demand_goods_order = ["食品", "服装", "手机"]

    for idx, folk_item in enumerate(folk_data["folks"]):
        r = ws_folk.max_row + 1
        name = FOLK_NAMES[idx] if idx < len(FOLK_NAMES) else f"群体{idx}"
        ws_folk.cell(row=r, column=1, value=name)
        ws_folk.cell(row=r, column=2, value=folk_item["population"])
        ws_folk.cell(row=r, column=3, value=folk_item["w_quality"])
        ws_folk.cell(row=r, column=4, value=folk_item["w_brand"])
        ws_folk.cell(row=r, column=5, value=folk_item["w_price"])
        ws_folk.cell(row=r, column=6, value=folk_item["labor_participation_rate"])
        ws_folk.cell(row=r, column=7, value=folk_item["labor_points_per_capita"])
        col = 8
        for gname in demand_goods_order:
            bd = folk_item.get("base_demands", {})
            demand_cfg = bd.get(gname, {"per_capita": 0, "sensitivity": 0})
            ws_folk.cell(row=r, column=col, value=demand_cfg["per_capita"])
            ws_folk.cell(row=r, column=col + 1, value=demand_cfg["sensitivity"])
            col += 2
        for c in range(1, len(folk_headers) + 1):
            _style_data(ws_folk.cell(row=r, column=c))
    _auto_width(ws_folk)

    # ═══════════ Sheet 5: 经济分析 ═══════════
    ws4 = wb.create_sheet("经济分析")
    wage_cell = "经济分析!$B$1"

    ws4.cell(row=1, column=1, value="统一工资").font = Font(bold=True)
    ws4.cell(row=1, column=2, value=10)
    _style_param(ws4["B1"])

    header_row = 3
    headers4 = [
        "工厂名称", "产业链", "产出商品", "产出数量", "商品售价",
        "收入", "劳动力需求", "工资支出", "维护费",
        "输入商品", "输入数量", "原料单价", "原料成本",
        "总成本", "利润", "利润率", "建造成本", "单回合ROI", "回本回合数",
    ]
    for c, h in enumerate(headers4, 1):
        ws4.cell(row=header_row, column=c, value=h)
    _style_header(ws4, header_row, len(headers4))

    data_start = header_row + 1
    for i, ft in enumerate(goods_data["factory_types"]):
        row = data_start + i
        recipe = recipe_map[ft["recipe"]]
        chain = CHAIN_MAP.get(ft["name"], "")
        output_name = recipe["output"]
        input_name = recipe.get("input")
        recipe_row = 2 + list(recipe_map.keys()).index(ft["recipe"])
        ft_row = 2 + i
        gp_row = goods_price_row[output_name]

        ws4.cell(row=row, column=1, value=ft["name"])
        ws4.cell(row=row, column=2, value=chain)
        ws4.cell(row=row, column=3, value=output_name)
        ws4.cell(row=row, column=4).value = f"=配方!E{recipe_row}"
        ws4.cell(row=row, column=5).value = f"=商品类型!B{gp_row}"
        ws4.cell(row=row, column=6).value = f"=D{row}*E{row}"
        ws4.cell(row=row, column=7).value = f"=工厂类型!C{ft_row}"
        ws4.cell(row=row, column=8).value = f"=G{row}*$B$1"
        ws4.cell(row=row, column=9).value = f"=工厂类型!E{ft_row}"
        ws4.cell(row=row, column=10, value=input_name or "(无)")
        ws4.cell(row=row, column=11).value = f"=配方!C{recipe_row}"
        if input_name:
            igp_row = goods_price_row[input_name]
            ws4.cell(row=row, column=12).value = f"=商品类型!B{igp_row}"
        else:
            ws4.cell(row=row, column=12, value=0)
        ws4.cell(row=row, column=13).value = f"=K{row}*L{row}"
        ws4.cell(row=row, column=14).value = f"=H{row}+I{row}+M{row}"
        ws4.cell(row=row, column=15).value = f"=F{row}-N{row}"
        ws4.cell(row=row, column=16).value = f"=IF(F{row}=0,0,O{row}/F{row})"
        ws4.cell(row=row, column=17).value = f"=工厂类型!D{ft_row}"
        ws4.cell(row=row, column=18).value = f"=IF(Q{row}=0,0,O{row}/Q{row})"
        ws4.cell(row=row, column=19).value = f'=IF(O{row}<=0,"亏损",Q{row}/O{row})'

        chain_fill = CHAIN_FILLS.get(chain)
        for c in range(1, len(headers4) + 1):
            cell = ws4.cell(row=row, column=c)
            is_f = isinstance(cell.value, str) and str(cell.value).startswith("=")
            _style_data(cell, is_formula=is_f)
            if chain_fill and not is_f:
                cell.fill = chain_fill
        ws4.cell(row=row, column=16).number_format = '0.0%'
        ws4.cell(row=row, column=18).number_format = '0.0%'
        ws4.cell(row=row, column=19).number_format = '0.0'

    _auto_width(ws4)
    ws4.freeze_panes = "A4"

    # ═══════════ Sheet 6: 居民需求分析 ═══════════
    ws_demand = wb.create_sheet("需求分析")

    # --- 经济周期参数区 ---
    a1 = economy_data["dual_cycle"]["short_cycle"]["amplitude"]
    a2 = economy_data["dual_cycle"]["long_cycle"]["amplitude"]
    max_index = min(a1 + a2, 1.0)
    min_index = max(-(a1 + a2), -1.0)

    ws_demand.cell(row=1, column=1, value="经济周期参数").font = SECTION_FONT
    ws_demand.cell(row=2, column=1, value="短周期振幅 A1")
    ws_demand.cell(row=2, column=2, value=a1)
    ws_demand.cell(row=3, column=1, value="长周期振幅 A2")
    ws_demand.cell(row=3, column=2, value=a2)
    ws_demand.cell(row=4, column=1, value="理论最高指数")
    ws_demand.cell(row=4, column=2, value=max_index)
    ws_demand.cell(row=5, column=1, value="理论最低指数")
    ws_demand.cell(row=5, column=2, value=min_index)
    for r in range(2, 6):
        for c in (1, 2):
            _style_data(ws_demand.cell(row=r, column=c))

    # --- 需求公式: demand = population * per_capita * (1 + index * sensitivity) ---
    # 居民配置表引用: 行2=低收入, 行3=中等, 行4=高收入
    # 居民配置列: B=人口, H=食品人均, I=食品敏感, J=服装人均, K=服装敏感, L=手机人均, M=手机敏感
    folk_pop_cols = ["B"]  # 人口列
    # 每种商品在居民配置表中的(人均列, 敏感度列)
    goods_folk_cols = {
        "食品": ("H", "I"),
        "服装": ("J", "K"),
        "手机": ("L", "M"),
    }
    folk_rows_in_sheet = [2, 3, 4]  # 居民配置表中的数据行

    # --- 三个场景 ---
    scenarios = [
        ("经济指数 = 0 (基准)", 0),
        (f"经济指数 = {max_index} (繁荣顶峰)", max_index),
        (f"经济指数 = {min_index} (衰退谷底)", min_index),
    ]

    cur_row = 7

    # 终端商品 → 产业链工厂信息 (用于后续计算需要多少工厂)
    # 每个终端商品的单工厂产出 = 配方表中终端配方的 output_quantity
    # 还需要追溯上游: 终端需要多少中间品 → 需要多少中间工厂 → 需要多少原料 → 需要多少原料工厂
    # 产业链结构: goods_data["recipes"] 排列为 [原料,中间,终端] 每3个一组
    chain_info = {}  # {终端商品名: {level: {recipe_name, output_qty, input_qty, ...}}}
    chains_order = [
        ("手机", ["硅矿开采", "芯片制造", "手机组装"], ["硅矿场", "芯片工厂", "手机工厂"]),
        ("服装", ["棉花种植", "布料加工", "服装生产"], ["棉花农场", "纺织厂", "服装厂"]),
        ("食品", ["小麦种植", "面粉加工", "食品生产"], ["麦田", "面粉厂", "食品厂"]),
    ]

    for s_label, s_index in scenarios:
        _style_section(ws_demand, cur_row, 10, s_label)
        cur_row += 1

        # 子表头
        sub_headers = ["商品", "群体", "人口", "人均需求", "敏感度", "需求量"]
        for c, h in enumerate(sub_headers, 1):
            ws_demand.cell(row=cur_row, column=c, value=h)
        _style_header(ws_demand, cur_row, len(sub_headers))
        cur_row += 1

        # 记录每种商品的需求总量行，供后续汇总引用
        goods_demand_rows = {}  # {商品名: [各群体需求行号]}

        for gname in demand_goods_order:
            per_cap_col, sens_col = goods_folk_cols[gname]
            row_list = []
            for fi, fr in enumerate(folk_rows_in_sheet):
                ws_demand.cell(row=cur_row, column=1, value=gname if fi == 0 else "")
                ws_demand.cell(row=cur_row, column=2, value=FOLK_NAMES[fi])
                # 人口 → 引用居民配置
                ws_demand.cell(row=cur_row, column=3).value = f"=居民配置!B{fr}"
                # 人均需求
                ws_demand.cell(row=cur_row, column=4).value = f"=居民配置!{per_cap_col}{fr}"
                # 敏感度
                ws_demand.cell(row=cur_row, column=5).value = f"=居民配置!{sens_col}{fr}"
                # 需求量 = INT(人口 * 人均 * (1 + index * 敏感度))
                ws_demand.cell(row=cur_row, column=6).value = (
                    f"=INT(C{cur_row}*D{cur_row}*(1+{s_index}*E{cur_row}))"
                )
                for c in range(1, 7):
                    _style_data(ws_demand.cell(row=cur_row, column=c), is_formula=(c >= 3))
                row_list.append(cur_row)
                cur_row += 1

            # 小计行
            ws_demand.cell(row=cur_row, column=1, value="")
            ws_demand.cell(row=cur_row, column=2, value=f"{gname} 合计")
            ws_demand.cell(row=cur_row, column=2).font = Font(bold=True)
            sum_refs = "+".join(f"F{r}" for r in row_list)
            ws_demand.cell(row=cur_row, column=6).value = f"={sum_refs}"
            ws_demand.cell(row=cur_row, column=6).font = Font(bold=True)
            for c in range(1, 7):
                _style_data(ws_demand.cell(row=cur_row, column=c), is_formula=(c == 6))
            goods_demand_rows[gname] = cur_row  # 记录合计行
            cur_row += 1

        # --- 需要多少工厂 ---
        cur_row += 1
        _style_section(ws_demand, cur_row, 10, f"所需工厂数量 ({s_label})")
        cur_row += 1
        factory_headers = ["产业链", "层级", "工厂类型", "单厂产出", "终端需求", "原料需求系数", "所需产量", "所需工厂数(向上取整)"]
        for c, h in enumerate(factory_headers, 1):
            ws_demand.cell(row=cur_row, column=c, value=h)
        _style_header(ws_demand, cur_row, len(factory_headers))
        cur_row += 1

        for end_good, recipe_names, factory_names in chains_order:
            chain_name = CHAIN_MAP[end_good]
            demand_sum_row = goods_demand_rows[end_good]
            # 终端层
            end_recipe = recipe_map[recipe_names[2]]
            mid_recipe = recipe_map[recipe_names[1]]
            raw_recipe = recipe_map[recipe_names[0]]

            # 引用配方表行号
            end_recipe_row = 2 + list(recipe_map.keys()).index(recipe_names[2])
            mid_recipe_row = 2 + list(recipe_map.keys()).index(recipe_names[1])
            raw_recipe_row = 2 + list(recipe_map.keys()).index(recipe_names[0])

            # 终端工厂行
            r = cur_row
            ws_demand.cell(row=r, column=1, value=chain_name)
            ws_demand.cell(row=r, column=2, value="终端")
            ws_demand.cell(row=r, column=3, value=factory_names[2])
            ws_demand.cell(row=r, column=4).value = f"=配方!E{end_recipe_row}"  # 单厂产出
            ws_demand.cell(row=r, column=5).value = f"=F{demand_sum_row}"  # 终端需求
            ws_demand.cell(row=r, column=6, value=1)  # 系数=1
            ws_demand.cell(row=r, column=7).value = f"=E{r}"  # 所需产量=终端需求
            ws_demand.cell(row=r, column=8).value = f"=ROUNDUP(G{r}/D{r},0)"
            for c in range(1, 9):
                _style_data(ws_demand.cell(row=r, column=c), is_formula=(c >= 4))
            end_row = r
            cur_row += 1

            # 中间工厂行
            # 所需中间品数量 = 终端工厂数 * 终端配方input_quantity
            # 中间工厂所需产量 = ceil(终端需求 / 终端单产) * 终端input_quantity
            r = cur_row
            ws_demand.cell(row=r, column=1, value="")
            ws_demand.cell(row=r, column=2, value="中间")
            ws_demand.cell(row=r, column=3, value=factory_names[1])
            ws_demand.cell(row=r, column=4).value = f"=配方!E{mid_recipe_row}"
            ws_demand.cell(row=r, column=5).value = f"=E{end_row}"
            # 系数 = 终端input_qty / 终端output_qty
            ws_demand.cell(row=r, column=6).value = f"=配方!C{end_recipe_row}/配方!E{end_recipe_row}"
            ws_demand.cell(row=r, column=7).value = f"=ROUNDUP(E{r}*F{r},0)"  # 所需中间品产量
            ws_demand.cell(row=r, column=8).value = f"=ROUNDUP(G{r}/D{r},0)"
            for c in range(1, 9):
                _style_data(ws_demand.cell(row=r, column=c), is_formula=(c >= 4))
            mid_row = r
            cur_row += 1

            # 原料工厂行
            r = cur_row
            ws_demand.cell(row=r, column=1, value="")
            ws_demand.cell(row=r, column=2, value="原料")
            ws_demand.cell(row=r, column=3, value=factory_names[0])
            ws_demand.cell(row=r, column=4).value = f"=配方!E{raw_recipe_row}"
            ws_demand.cell(row=r, column=5).value = f"=E{end_row}"
            # 系数 = (终端input/终端output) * (中间input/中间output)
            ws_demand.cell(row=r, column=6).value = (
                f"=(配方!C{end_recipe_row}/配方!E{end_recipe_row})"
                f"*(配方!C{mid_recipe_row}/配方!E{mid_recipe_row})"
            )
            ws_demand.cell(row=r, column=7).value = f"=ROUNDUP(E{r}*F{r},0)"
            ws_demand.cell(row=r, column=8).value = f"=ROUNDUP(G{r}/D{r},0)"
            for c in range(1, 9):
                _style_data(ws_demand.cell(row=r, column=c), is_formula=(c >= 4))
            cur_row += 1

        cur_row += 1  # 场景间空行

    _auto_width(ws_demand)
    ws_demand.freeze_panes = "A2"

    # ═══════════ Sheet 7: 现金总量分析 ═══════════
    ws_cash = wb.create_sheet("现金总量分析")

    companies = game_data["companies"]

    # --- 区块1: 各部门现金明细 ---
    cur_row = 1
    _style_section(ws_cash, cur_row, 5, "各部门初始现金明细")
    cur_row += 1
    cash_headers = ["类别", "名称", "数量", "单位现金", "小计"]
    for c, h in enumerate(cash_headers, 1):
        ws_cash.cell(row=cur_row, column=c, value=h)
    _style_header(ws_cash, cur_row, len(cash_headers))
    cur_row += 1

    company_start = cur_row
    for comp in companies:
        ws_cash.cell(row=cur_row, column=1, value="企业")
        ws_cash.cell(row=cur_row, column=2, value=comp["factory_type"])
        ws_cash.cell(row=cur_row, column=3, value=comp["count"])
        ws_cash.cell(row=cur_row, column=4, value=comp["initial_cash"])
        ws_cash.cell(row=cur_row, column=5).value = f"=C{cur_row}*D{cur_row}"
        for c in range(1, 6):
            _style_data(ws_cash.cell(row=cur_row, column=c), is_formula=(c == 5))
        cur_row += 1
    company_end = cur_row - 1

    # 企业小计
    ws_cash.cell(row=cur_row, column=1, value="")
    ws_cash.cell(row=cur_row, column=2, value="企业合计").font = Font(bold=True)
    ws_cash.cell(row=cur_row, column=5).value = f"=SUM(E{company_start}:E{company_end})"
    ws_cash.cell(row=cur_row, column=5).font = Font(bold=True)
    for c in range(1, 6):
        _style_data(ws_cash.cell(row=cur_row, column=c), is_formula=(c == 5))
    company_sum_row = cur_row
    cur_row += 1

    # 银行
    bank_start = cur_row
    for bank in game_data["banks"]:
        ws_cash.cell(row=cur_row, column=1, value="银行")
        ws_cash.cell(row=cur_row, column=2, value=bank["name"])
        ws_cash.cell(row=cur_row, column=3, value=1)
        ws_cash.cell(row=cur_row, column=4, value=bank["initial_cash"])
        ws_cash.cell(row=cur_row, column=5).value = f"=C{cur_row}*D{cur_row}"
        for c in range(1, 6):
            _style_data(ws_cash.cell(row=cur_row, column=c), is_formula=(c == 5))
        cur_row += 1
    bank_end = cur_row - 1

    ws_cash.cell(row=cur_row, column=2, value="银行合计").font = Font(bold=True)
    ws_cash.cell(row=cur_row, column=5).value = f"=SUM(E{bank_start}:E{bank_end})"
    ws_cash.cell(row=cur_row, column=5).font = Font(bold=True)
    for c in range(1, 6):
        _style_data(ws_cash.cell(row=cur_row, column=c), is_formula=(c == 5))
    bank_sum_row = cur_row
    cur_row += 1

    # 居民
    folk_count = len(folk_data["folks"])
    ws_cash.cell(row=cur_row, column=1, value="居民")
    ws_cash.cell(row=cur_row, column=2, value=f"全体居民({folk_count}群体)")
    ws_cash.cell(row=cur_row, column=3, value=folk_count)
    ws_cash.cell(row=cur_row, column=4, value=game_data["folk_initial_cash"])
    ws_cash.cell(row=cur_row, column=5).value = f"=C{cur_row}*D{cur_row}"
    for c in range(1, 6):
        _style_data(ws_cash.cell(row=cur_row, column=c), is_formula=(c == 5))
    folk_sum_row = cur_row
    cur_row += 1

    # 总计
    cur_row += 1
    ws_cash.cell(row=cur_row, column=2, value="系统现金总量").font = Font(bold=True, size=12)
    ws_cash.cell(row=cur_row, column=5).value = f"=E{company_sum_row}+E{bank_sum_row}+E{folk_sum_row}"
    ws_cash.cell(row=cur_row, column=5).font = Font(bold=True, size=12)
    ws_cash.cell(row=cur_row, column=5).number_format = '#,##0'
    for c in range(1, 6):
        _style_data(ws_cash.cell(row=cur_row, column=c), is_formula=(c == 5))
    total_cash_row = cur_row
    cur_row += 2

    # --- 区块2: 现金占比 ---
    _style_section(ws_cash, cur_row, 5, "现金分布占比")
    cur_row += 1
    ratio_headers = ["部门", "现金", "占比"]
    for c, h in enumerate(ratio_headers, 1):
        ws_cash.cell(row=cur_row, column=c, value=h)
    _style_header(ws_cash, cur_row, len(ratio_headers))
    cur_row += 1

    for label, sum_row in [("企业", company_sum_row), ("银行", bank_sum_row), ("居民", folk_sum_row)]:
        ws_cash.cell(row=cur_row, column=1, value=label)
        ws_cash.cell(row=cur_row, column=2).value = f"=E{sum_row}"
        ws_cash.cell(row=cur_row, column=2).number_format = '#,##0'
        ws_cash.cell(row=cur_row, column=3).value = f"=E{sum_row}/E{total_cash_row}"
        ws_cash.cell(row=cur_row, column=3).number_format = '0.0%'
        for c in range(1, 4):
            _style_data(ws_cash.cell(row=cur_row, column=c), is_formula=(c >= 2))
        cur_row += 1

    cur_row += 1

    # --- 区块3: 现金 vs 工厂建设价格 ---
    _style_section(ws_cash, cur_row, 5, "现金总量 / 工厂建造成本")
    cur_row += 1
    build_headers = ["工厂类型", "建造成本", "现金总量", "可建造数量", "备注"]
    for c, h in enumerate(build_headers, 1):
        ws_cash.cell(row=cur_row, column=c, value=h)
    _style_header(ws_cash, cur_row, len(build_headers))
    cur_row += 1

    for i, ft in enumerate(goods_data["factory_types"]):
        ft_row = 2 + i  # 工厂类型表中的行
        ws_cash.cell(row=cur_row, column=1, value=ft["name"])
        ws_cash.cell(row=cur_row, column=2).value = f"=工厂类型!D{ft_row}"
        ws_cash.cell(row=cur_row, column=2).number_format = '#,##0'
        ws_cash.cell(row=cur_row, column=3).value = f"=E{total_cash_row}"
        ws_cash.cell(row=cur_row, column=3).number_format = '#,##0'
        ws_cash.cell(row=cur_row, column=4).value = f"=INT(C{cur_row}/B{cur_row})"
        ws_cash.cell(row=cur_row, column=5).value = (
            f'=IF(D{cur_row}<20,"偏贵",IF(D{cur_row}>200,"偏便宜","合理"))'
        )
        chain = CHAIN_MAP.get(ft["name"])
        for c in range(1, 6):
            cell = ws_cash.cell(row=cur_row, column=c)
            _style_data(cell, is_formula=(c >= 2))
            if chain in CHAIN_FILLS and c <= 1:
                cell.fill = CHAIN_FILLS[chain]
        cur_row += 1

    _auto_width(ws_cash)

    wb.save(EXCEL_PATH)
    print(f"已导出: {EXCEL_PATH}")


# ────────────────── IMPORT ──────────────────


def import_from_excel():
    wb = load_workbook(EXCEL_PATH, data_only=True)

    # 读取商品类型
    ws1 = wb["商品类型"]
    goods_types = []
    for row in ws1.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            break
        goods_types.append({"name": row[0], "base_price": int(row[1])})

    # 读取配方
    ws2 = wb["配方"]
    recipes = []
    for row in ws2.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            break
        recipes.append({
            "name": row[0],
            "input": None if row[1] == "(无)" else row[1],
            "input_quantity": int(row[2]),
            "output": row[3],
            "output_quantity": int(row[4]),
            "tech_quality_weight": float(row[5]),
        })

    # 读取工厂类型
    ws3 = wb["工厂类型"]
    factory_types = []
    for row in ws3.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            break
        factory_types.append({
            "name": row[0],
            "recipe": row[1],
            "labor_demand": int(row[2]),
            "build_cost": int(row[3]),
            "maintenance_cost": int(row[4]),
            "build_time": int(row[5]),
        })

    goods_data = {
        "goods_types": goods_types,
        "recipes": recipes,
        "factory_types": factory_types,
    }
    _save_yaml(goods_data, GOODS_YAML)
    print(f"已导入: {GOODS_YAML}")

    # 读取居民配置
    ws_folk = wb["居民配置"]
    demand_goods_order = ["食品", "服装", "手机"]
    original_folk = _load_yaml(FOLK_YAML)

    folks = []
    for row_idx, row in enumerate(ws_folk.iter_rows(min_row=2, values_only=True)):
        if row[0] is None:
            break
        # 保留原始 spending_flow
        orig = original_folk["folks"][row_idx] if row_idx < len(original_folk["folks"]) else {}
        spending_flow = orig.get("spending_flow", {"tech": 0.3, "brand": 0.4, "maintenance": 0.3})

        base_demands = {}
        col = 7  # 0-indexed: H=7, I=8, J=9, K=10, L=11, M=12
        for gname in demand_goods_order:
            per_cap = row[col]
            sens = row[col + 1]
            base_demands[gname] = {"per_capita": per_cap, "sensitivity": sens}
            col += 2

        folk_entry = {
            "population": int(row[1]),
            "w_quality": float(row[2]),
            "w_brand": float(row[3]),
            "w_price": float(row[4]),
            "labor_participation_rate": float(row[5]),
            "labor_points_per_capita": float(row[6]),
            "spending_flow": spending_flow,
            "base_demands": base_demands,
        }
        folks.append(folk_entry)

    folk_data = {"folks": folks}
    _save_yaml(folk_data, FOLK_YAML)
    print(f"已导入: {FOLK_YAML}")


# ────────────────── main ──────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="游戏配置 <-> Excel 转换工具")
    parser.add_argument("action", choices=["export", "import"],
                        help="export=YAML→Excel, import=Excel→YAML")
    args = parser.parse_args()

    if args.action == "export":
        export_to_excel()
    else:
        import_from_excel()
